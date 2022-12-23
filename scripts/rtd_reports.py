import os
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
import warnings
import re
import shutil

import utils.date_time as dt
from utils.helpers import move_file
from utils.score_card import handle_scorecard


def filter_reports():
    report_directory = r'Z:\rtd_reports'

    sub_dir = 'data'
    output_directory = os.path.join(report_directory, sub_dir)

    if os.path.exists(os.path.join(report_directory, sub_dir)) == False:
        os.mkdir(os.path.join(report_directory, sub_dir))

    for filename in os.listdir(report_directory):
        # define filename
        f = os.path.join(report_directory, filename)
        # print(f)
        # check if file exists
        # include_files = ['EDM_02_', 'EDM_03_', 'EDM_04_', 'EDM_06_', 'EDM_07_', 'EDM_08_', 'EDM_09_', 'EDM_11_',
        #                  'EDM_12_', 'EDM_13_', 'EDM_14_', 'EDM_15_', 'EDM_17_', 'EDM_24_', 'EDM_27_', 'EDM_28_', 'EDM_29_', 'EDM_30_']
        if os.path.isfile(f):
            if 'EDM_02_' in f:
                move_file(f, output_directory)
            if 'EDM_03_' in f:
                move_file(f, output_directory)
            if 'EDM_04_' in f:
                move_file(f, output_directory)
            if 'EDM_06_' in f:
                move_file(f, output_directory)
            if 'EDM_07_' in f:
                move_file(f, output_directory)
            if 'EDM_08_' in f:
                move_file(f, output_directory)
            if 'EDM_09_' in f:
                move_file(f, output_directory)
            if 'EDM_11_' in f:
                move_file(f, output_directory)
            if 'EDM_12_' in f:
                move_file(f, output_directory)
            if 'EDM_13_' in f:
                move_file(f, output_directory)
            if 'EDM_14_' in f:
                move_file(f, output_directory)
            if 'EDM_15_' in f:
                move_file(f, output_directory)
            if 'EDM_17_' in f:
                move_file(f, output_directory)
            if 'EDM_24_' in f:
                move_file(f, output_directory)
            if 'EDM_27_' in f:
                move_file(f, output_directory)
            if 'EDM_28_' in f:
                move_file(f, output_directory)
            if 'EDM_29_' in f:
                move_file(f, output_directory)
            if 'EDM_30_' in f:
                move_file(f, output_directory)

    num_files = len([f for f in os.listdir(output_directory)
                    if os.path.isfile(os.path.join(output_directory, f))])
    input("Reports moved to DATA directory:  %s. \nPress ENTER key to finish. " % (num_files))


def convert_reports():
    warnings.simplefilter("ignore")

    # PRODUCTION DIRECTORY
    # report_directory = "Z:\\rtd_reports"
    report_directory = (
        r"C:\RA-Apps\LT-Sync"
    )
    # DEV REPORT DIRECTORY
    # report_directory = os.path.join(os.getcwd())
    # print(f'WORKING DIRECTORY: \n{report_directory}\n')

    report_date = input("Please provide report date in MM-DD-YYYY format: \n")

    print()

    scorecard = {}

    data_directory = os.path.join(report_directory,  'INPUTS', "data")
    output_directory = os.path.join(report_directory, report_date)

    if os.path.exists(os.path.join(report_directory, report_date)) == False:
        os.mkdir(os.path.join(report_directory, report_date))

    for filename in os.listdir(data_directory):
        # define filename
        f = os.path.join(data_directory, filename)
        # chekc if file exists
        if os.path.isfile(f):
            # read data
            df = pd.read_csv(f, sep="\t")
            # rename output dir in filname
            if "EDM_02_Make_wo_LT_" in f:
                # TODO: add relevat inforamtion to scoracard DF
                new_name = os.path.join(
                    output_directory, f"{report_date} Make wo LT.xlsx")
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["Plant"], values=["Material"], aggfunc=["count"], margins="true"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_03_sto_wo_lt_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} STO_wo_LT.xlsx")
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["Plant"], values=["Material"], aggfunc=["count"], margins="true"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_04_tot_rep_LT_vs_IPT_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} Tot Rep LT vs IPT.xlsx"
                )
                scorecard["rep_lt_isu"] = df.shape[0]
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["Plant"], values=["Material"], aggfunc=["count"], margins="true"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_06_dist_center_make_parts_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} Distribution Make parts.xlsx"
                )
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["Plant"], values=["Material"], aggfunc=["count"], margins="true"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_07_STO_ND_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} STO ND.xlsx")
                scorecard["plant_nd_isu"] = df.shape[0]
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["SupPlnt"],
                    values=["Material"],
                    aggfunc=["count"],
                    margins="true",
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_08_dist_center_make_parts_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} Distribution Make parts.xlsx"
                )
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["Plant"], values=["Material"], aggfunc=["count"], margins="true"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_08_Parts_planning_strgr_consistency_check_" in f:
                new_name = os.path.join(
                    output_directory,
                    f"{report_date} Parts_Planning_Strategy_Consistency_Check.xlsx",
                )
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["Werks"], values=["Matnr"], aggfunc=["count"], margins="true"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_09_not_extended_to_SPK_plant_" in f:
                scorecard["spk_plant_isu"] = df.shape[0]
                new_name = os.path.join(
                    output_directory, f"{report_date} SPK not extended to plant.xlsx"
                )
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["RECV_PLANT"],
                    values=["Material"],
                    aggfunc=["count"],
                    margins="true",
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_11_Strat_group_check_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} Strategy_Grp_Check.xlsx"
                )
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["Werks"], values=["Matnr"], aggfunc=["count"], margins="true"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_12_Matl_Proc_to_Spec_Proc_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} Matl Proc to Spec Proc.xlsx"
                )
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["Plant"], values=["Material"], aggfunc=["count"], margins="true"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_13_min_max_lots_equal_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} Min Max Lots equal.xlsx"
                )
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["Plant"], values=["Material"], aggfunc=["count"], margins="true"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_13_min_max_lots_equal_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} Min Max Lots equal.xlsx"
                )
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["Plant"], values=["Material"], aggfunc=["count"], margins="true"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_14_Purch_no_EP_sloc_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} Purch no EP sloc.xlsx"
                )
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["Werks"], values=["Matnr"], aggfunc=["count"], margins="true"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_15_STO_parts_wo_released_std_cost_" in f:
                scorecard["std_cost_isu"] = df.shape[0]
                new_name = os.path.join(
                    output_directory, f"{report_date} STO_Parts_wo_Released_Std_Cost.xlsx"
                )
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["RcvPlnt"],
                    values=["Material"],
                    aggfunc=["count"],
                    margins="true",
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_17_Quota_Arranement_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} Quota_arrangement.xlsx"
                )
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["Plant"], values=["Material"], aggfunc=["count"], margins="true"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)
            if "EDM_24_SPK_VS_DEL_PLANT_CHECK_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} del_plant_validation.xlsx"
                )
                # CREATE PIVOT
                pivot = df.pivot_table(
                    index=["PLANTID"],
                    values=["MATERIALID"],
                    aggfunc=["count"],
                    margins="true",
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    pivot.to_excel(writer, sheet_name="summary")
                    df.to_excel(writer, sheet_name="DATA", index=False)
                # log filename
                print(new_name)

            if "EDM_27_NOT_EXT_SPK_PLANT_COUNT_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} Not Extended to SPK Plant Count.xlsx"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    df.to_excel(writer, sheet_name="Sheet1", index=False)
                # log filename
                print(new_name)
            if "EDM_28_STO_FROM_PLANT_WITH_ND_COUNT_" in f:
                new_name = os.path.join(
                    output_directory,
                    f"{report_date} STO From Plant with NO Planning_ND_Count.xlsx",
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    df.to_excel(writer, sheet_name="Sheet1", index=False)
                # log filename
                print(new_name)
            if "EDM_29_STO_NO_COST_COUNT_" in f:
                new_name = os.path.join(
                    output_directory,
                    f"{report_date} Total Rep Lead Time vs In House Prod Time Count.xlsx",
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    df.to_excel(writer, sheet_name="Sheet1", index=False)
                # log filename
                print(new_name)
            if "EDM_30_TRLT_VS_IPT_COUNT_" in f:
                new_name = os.path.join(
                    output_directory, f"{report_date} STO No Cost Count.xlsx"
                )
                # output data farame without df index & Pivot
                with pd.ExcelWriter(os.path.join(output_directory, new_name)) as writer:
                    df.to_excel(writer, sheet_name="Sheet1", index=False)
                # log filename
                print(new_name)
    print("\nRTD reports formated and edited\n")

    # TODO: save Scorecard DF

    # SUM VALUES FOR COUNT REPORTS V.1
    # for filename in os.listdir(os.path.join(report_directory, report_date)):
    #   # define filename
    #   f = os.path.join(report_directory, report_date , filename)
    #   # chekc if file exists
    #   if os.path.isfile(f):
    #     if 'Count' in f:
    #       wb = load_workbook(filename=f)
    #       sht = wb.active
    #       sht['D2'] = '=sum(B2:B200)'
    #       wb.save(filename=f)
    #       print(f'{f} - file edited')

    # SUM VALUES FOR COUNT REPORTS V.2
    for filename in os.listdir(os.path.join(report_directory, report_date)):
        # define filename
        f = os.path.join(report_directory, report_date, filename)
        # chekc if file exists
        if os.path.isfile(f):
            if "Count" in f:
                df = pd.read_excel(f, sheet_name="Sheet1")
                total = df["QTY"].sum()
                print(f"{filename}: {total}")
                if "Total Rep Lead Time" in filename:
                    scorecard["rep_lt_pop"] = total
                if "NO Planning_ND_Count" in filename:
                    scorecard["plant_nd_pop"] = total
                if "SPK Plant Count" in filename:
                    scorecard["spk_plant_pop"] = total
                if "No Cost Count" in filename:
                    scorecard["std_cost_pop"] = total

                wb = load_workbook(filename=f)
                sht = wb.active
                sht["D2"] = total
                wb.save(filename=f)
                print(f"{f} - count file sum edited")

    input(
        f"\nFinished! Please find reports in {report_date} directory. \nPress ENTER to close."
    )

    # SCORECARD DATA
    handle_scorecard(scorecard)


def archive_reports():
    report_directory = 'Z:\\rtd_reports'

    data_directory = os.path.join(report_directory, 'data')
    archive_directory = os.path.join(
        report_directory, 'ARCHIVE-PLANNING_PLAUSE')

    # print(data_directory)
    # print(archive_directory)

    # archive data folder with rtd reports
    for filename in os.listdir(data_directory):
        f = os.path.join(data_directory, filename)
        if os.path.isfile(f):
            shutil.move(f, archive_directory)

    # remove tmp dirs: data, input
    if os.path.isdir(report_directory):
        shutil.rmtree(data_directory)

    # find and remove output dire
    for filename in os.listdir(report_directory):
        match = re.search(r"^\d{2}\-\d{2}\-\d{4}", filename)
        if match:
            if os.path.isdir(os.path.join(report_directory, filename)):
                shutil.rmtree(os.path.join(report_directory, filename))
